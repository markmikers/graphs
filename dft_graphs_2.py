import numpy as np
import matplotlib.pyplot as plt
import matplotlib.lines as ml
import matplotlib.patches as mp
import openpyxl
import re
import copy

wb = openpyxl.load_workbook('./binaries_fcc_database14_reconsidered_vo.xlsx')
rootdir = './dft_graphs'



equilibrium_volumes = {'Cr': [11.75, 11.41], 'Mn': [11.3725, 10.7387931], 'Fe': [11.19625, 11.35], 'Ni': [10.91, 10.91]}

def plot_emix():
    # Enthalpy of mixing
    same_atomic_ratio = [[], [], [], [], [], [], [], [], []]
    for i in range(N):
        j = concentrations.index(colnames['02concentration'][i])
        k = colnames['01names'][i]
        same_atomic_ratio[j].append(k)
    x_gs = gs_colnames['02concentration']
    y_gs = gs_colnames['08emix']
    x_ngs = ngs_colnames['02concentration']
    y_ngs = ngs_colnames['08emix']

    if min(colnames['08emix']) < 0:
        plt.axhline(y=0, xmin=0, xmax=1, linewidth=1, color='lightgray', zorder=1)
    # ch = [[0, 0, -1], [1, 0, 0]]
    #     for i in range(len(same_atomic_ratio)):
    #         for j in range(len(same_atomic_ratio[i])):
    #             same_atomic_ratio[i][j] = colnames['08emix'][colnames['01names'].index(same_atomic_ratio[i][j])]
    #     for i in same_atomic_ratio:
    #         ch.append([concentrations[same_atomic_ratio.index(i)], min(i)])
    #     ch.sort()
    #     for i in range(len(ch)-1):
    #         x1 = ch[i][0]
    #         x2 = ch[i+1][0]
    #         y1 = ch[i][1]
    #         y2 = ch[i+1][1]
    #
    #         try:
    #             tangent = (y2-y1)/(x2-x1)
    #         except ZeroDivisionError:
    #             continue
    #         ch[i+1].append(tangent)
    #     for k in range(len(ch)):
    #         for i in range(len(ch)-1):
    #             if ch[i+1][2] < ch[i][2]:
    #                 if (ch[i][0] != 0):
    #                     ch.pop(i)
    #                     break
    #
    #     xgs = []
    #     ygs = []
    #     for i in ch:
    #         xgs.append(i[0])
    #         ygs.append(i[1])
    #     plt.plot(xgs, ygs, 'g--', color='darkgrey', linewidth=1, zorder=1)

    plt.xlabel('Concentration of %s' % str(ws)[-4:-2])
    plt.ylabel('Enthalpy of mixing, eV')
    for _s, _x, _y in zip(markers, x_gs, y_gs):
        plt.scatter(_x, _y, marker=_s, edgecolors='darkorange', s=50, zorder=3, facecolors='none')
    for _s, _x, _y in zip(markers, x_ngs, y_ngs):
        plt.scatter(_x, _y, marker=_s, edgecolors='green', s=50, zorder=2, facecolors='none')
    plt.xlim(0, 1)
    filename = '%s/%s_emix' % (rootdir, str(ws)[-6:-2])
    # plt.savefig(filename)
    plt.show()
    plt.clf()


def plot_eform():
    # Enthalpy of formation
    same_atomic_ratio = [[], [], [], [], [], [], [], [], []]
    for i in range(N):
        j = concentrations.index(colnames['02concentration'][i])
        k = colnames['01names'][i]
        same_atomic_ratio[j].append(k)
    x_gs = gs_colnames['02concentration']
    y_gs = gs_colnames['09eform']
    x_ngs = ngs_colnames['02concentration']
    y_ngs = ngs_colnames['09eform']

    if min(colnames['09eform']) < 0:
        plt.axhline(y=0, xmin=0, xmax=1, linewidth=1, color='lightgray', zorder=1)
    # ch = [[0, 0, -1], [1, 0, 0]]
    #     for i in range(len(same_atomic_ratio)):
    #         for j in range(len(same_atomic_ratio[i])):
    #             same_atomic_ratio[i][j] = colnames['09eform'][colnames['01names'].index(same_atomic_ratio[i][j])]
    #     for i in same_atomic_ratio:
    #         ch.append([concentrations[same_atomic_ratio.index(i)], min(i)])
    #     ch.sort()
    #     for i in range(len(ch)-1):
    #         x1 = ch[i][0]
    #         x2 = ch[i+1][0]
    #         y1 = ch[i][1]
    #         y2 = ch[i+1][1]
    #
    #         try:
    #             tangent = (y2-y1)/(x2-x1)
    #         except ZeroDivisionError:
    #             continue
    #         ch[i+1].append(tangent)
    #     for k in range(len(ch)):
    #         for i in range(len(ch)-1):
    #             if ch[i+1][2] < ch[i][2]:
    #                 if (ch[i][0] != 0):
    #                     ch.pop(i)
    #                     break
    #
    #     xgs = []
    #     ygs = []
    #     for i in ch:
    #         xgs.append(i[0])
    #         ygs.append(i[1])
    #     plt.plot(xgs, ygs, 'g--', color='darkgrey', linewidth=1, zorder=1)

    plt.xlabel('Concentration of %s' % str(ws)[-4:-2])
    plt.ylabel('Enthalpy of formation, eV')
    for _s, _x, _y in zip(markers, x_gs, y_gs):
        plt.scatter(_x, _y, marker=_s, edgecolors='darkorange', s=50, zorder=3, facecolors='none')
    for _s, _x, _y in zip(markers, x_ngs, y_ngs):
        plt.scatter(_x, _y, marker=_s, edgecolors='green', s=50, zorder=2, facecolors='none')
    plt.xlim(0, 1)
    filename = '%s/%s_eform' % (rootdir, str(ws)[-6:-2])
    # plt.savefig(filename)
    plt.show()
    plt.clf()


def plot_volume():
    # Volume per atom
    same_atomic_ratio = [[], [], [], [], [], [], [], [], []]
    for i in range(N):
        j = concentrations.index(colnames['02concentration'][i])
        k = colnames['01names'][i]
        same_atomic_ratio[j].append(k)
    x_gs = gs_colnames['02concentration']
    # print len(x_gs)
    y_gs = gs_colnames['06v/at']
    x_ngs = ngs_colnames['02concentration']
    # print len(x_ngs)
    y_ngs = ngs_colnames['06v/at']
    plt.xlabel('Concentration of %s' % str(ws)[-4:-2])
    plt.ylabel('Volume per atom ' + r'$\.A^3$')
    for _s, _x, _y in zip(markers, x_gs, y_gs):
        plt.scatter(_x, _y, marker=_s, s=30, zorder=3, linewidths=2, edgecolors='darkorange', facecolors='none')
    for _s, _x, _y in zip(markers, x_ngs, y_ngs):
        plt.scatter(_x, _y, marker=_s, linewidths=1, edgecolors='green', s=50, zorder=2, facecolors='none')
    # print equilibrium_volumes[str(ws)[-4:-2]][0]
    # print equilibrium_volumes[str(ws)[-6:-4]][0]
    plt.plot((0, 1), (equilibrium_volumes[str(ws)[-6:-4]][0], equilibrium_volumes[str(ws)[-4:-2]][0]), linestyle='--', zorder=1)
    # plt.plot((1, 0), (equilibrium_volumes[str(ws)[-6:-4]][1], equilibrium_volumes[str(ws)[-4:-2]][1]), linestyle='--', zorder=1)
    plt.xlim(0, 1)
    veg_law = ml.Line2D([],[],linestyle='--', linewidth=1, color='blue', label="Vegard's law")
    gs = mp.Patch(color='darkorange', label='GS')
    ngs = mp.Patch(color='green', label='non-GS')
    fm = ml.Line2D([],[],linestyle='none', marker='^', markeredgecolor='blue', markerfacecolor='none', label='FM')
    afm = ml.Line2D([],[],linestyle='none', color='blue', marker='v', markeredgecolor='blue', markerfacecolor='none', label='AFM')
    fim = ml.Line2D([],[],linestyle='none', color='blue', marker='D', markeredgecolor='blue', markerfacecolor='none', label='FiM')
    nm = ml.Line2D([],[],linestyle='none', color='blue', marker='s', markeredgecolor='blue', markerfacecolor='none', label='NM')
    legend = plt.legend(handles=[nm, fm, afm, fim, veg_law, gs, ngs], bbox_to_anchor=(1.05, 1), loc=2, borderaxespad=0.)
    # plt.show()
    filename = '%s/%s_vol' % (rootdir, str(ws)[-6:-2])
    plt.savefig(filename, bbox_extra_artists=(legend,), bbox_inches='tight')
    plt.clf()


def plot_mmavg():
    # Average magnetic moment per atom
    same_atomic_ratio = [[], [], [], [], [], [], [], [], []]
    for i in range(N):
        j = concentrations.index(colnames['02concentration'][i])
        k = colnames['01names'][i]
        same_atomic_ratio[j].append(k)
    plt.axhline(y=0, xmin=0, xmax=1, linewidth=1, color='lightgray', zorder=1)
    x_gs = gs_colnames['02concentration']
    x_ngs = ngs_colnames['02concentration']
    y_gs = gs_colnames['11mmavg']
    y_ngs = ngs_colnames['11mmavg']
    plt.xlabel('Concentration of %s' % str(ws)[-4:-2])
    plt.ylabel('Average magnetic moment per atom, ' + r'$\mu_B$')
    for _s, _x, _y in zip(markers, x_gs, y_gs):
        plt.scatter(_x, _y, marker=_s, s=30, zorder=3, linewidths=2, edgecolors='darkorange', facecolors='none')
    for _s, _x, _y in zip(markers, x_ngs, y_ngs):
        plt.scatter(_x, _y, marker=_s, linewidths=1, edgecolors='green', s=50, zorder=2, facecolors='none')
    plt.xlim(0, 1)
    gs = mp.Patch(color='darkorange', label='GS')
    ngs = mp.Patch(color='green', label='non-GS')
    fm = ml.Line2D([], [], linestyle='none', marker='^', markeredgecolor='blue', markerfacecolor='none', label='FM')
    afm = ml.Line2D([], [], linestyle='none', color='blue', marker='v', markeredgecolor='blue', markerfacecolor='none',
                    label='AFM')
    fim = ml.Line2D([], [], linestyle='none', color='blue', marker='D', markeredgecolor='blue', markerfacecolor='none',
                    label='FiM')
    nm = ml.Line2D([], [], linestyle='none', color='blue', marker='s', markeredgecolor='blue', markerfacecolor='none',
                   label='NM')
    legend = plt.legend(handles=[nm, fm, afm, fim, gs, ngs], bbox_to_anchor=(1.05, 1), loc=2, borderaxespad=0.)
    filename = '%s/%s_mm' % (rootdir, str(ws)[-6:-2])
    plt.savefig(filename, bbox_extra_artists=(legend,), bbox_inches='tight')
    plt.clf()


def plot_mmsortavg():
    # Average magnetic moment per atomic sort
    same_atomic_ratio = [[], [], [], [], [], [], [], [], []]
    for i in range(N):
        j = concentrations.index(colnames['02concentration'][i])
        k = colnames['01names'][i]
        same_atomic_ratio[j].append(k)
    plt.axhline(y=0, xmin=0, xmax=1, linewidth=1, color='lightgray', zorder=1)
    x_gs = gs_colnames['02concentration']
    y1_gs = gs_colnames['12mm1avg']
    y2_gs = gs_colnames['13mm2avg']
    x_ngs = ngs_colnames['02concentration']
    y1_ngs = ngs_colnames['12mm1avg']
    y2_ngs = ngs_colnames['13mm2avg']
    plt.xlabel('Concentration of %s' % str(ws)[-4:-2])
    plt.ylabel('Average magnetic moment per atom, ' + r'$\mu_B$')
    for _x, _y in zip(x_gs, y1_gs):
        plt.scatter(_x, _y, marker='^', edgecolors='darkorange', s=100, zorder=3, facecolor='none')
    for _x, _y in zip(x_gs, y2_gs):
        plt.scatter(_x, _y, marker='v', edgecolors='darkorange', s=100, zorder=3, facecolor='none')
    for _x, _y in zip(x_ngs, y1_ngs):
        plt.scatter(_x, _y, marker='^', edgecolors='green', s=100, zorder=2, facecolor='none')
    for _x, _y in zip(x_ngs, y2_ngs):
        plt.scatter(_x, _y, marker='v', edgecolors='green', s=100, zorder=2, facecolor='none')
    plt.xlim(0, 1)
    at1 = ml.Line2D([], [], linestyle='none', color='blue', marker='^', markeredgecolor='blue', markerfacecolor='none',
                   label=str(ws)[-6:-4])
    at2 = ml.Line2D([], [], linestyle='none', color='blue', marker='v', markeredgecolor='blue', markerfacecolor='none',
                   label=str(ws)[-4:-2])
    gs = mp.Patch(color='darkorange', label='GS')
    ngs = mp.Patch(color='green', label='non-GS')
    legend = plt.legend(handles=[at1, at2, gs, ngs], bbox_to_anchor=(1.05, 1), loc=2, borderaxespad=0.)
    filename = '%s/%s_mmsort' % (rootdir, str(ws)[-6:-2])
    plt.savefig(filename, bbox_extra_artists=(legend,), bbox_inches='tight')
    # plt.show()
    plt.clf()


def plot_mmavg_and_mmsortavg():
    same_atomic_ratio = [[], [], [], [], [], [], [], [], []]
    for i in range(N):
        j = concentrations.index(colnames['02concentration'][i])
        k = colnames['01names'][i]
        same_atomic_ratio[j].append(k)
    plt.axhline(y=0, xmin=0, xmax=1, linewidth=1, color='lightgray', zorder=1)
    x = colnames['02concentration']
    y = colnames['11mmavg']
    y1 = colnames['12mm1avg']
    y2 = colnames['13mm2avg']
    plt.xlabel('Concentration of %s' % str(ws)[-4:-2])
    plt.ylabel('Average magnetic moment per atom, muB')
    for c, _x, _y, _bw in zip(fill_colors, x, y, border_width):
        plt.scatter(_x, _y, marker='o', c=c, linewidths=_bw, edgecolors='black', s=50, zorder=2)
    for c, _x, _y, _bw in zip(fill_colors, x, y1, border_width):
        plt.scatter(_x, _y, marker='^', c=c, linewidths=_bw, edgecolors='black', s=100, zorder=2)
    for c, _x, _y, _bw in zip(fill_colors, x, y2, border_width):
        plt.scatter(_x, _y, marker='v', c=c, linewidths=_bw, edgecolors='black', s=100, zorder=2)
    plt.xlim(0, 1)
    filename = '%s/%s_mmavg_and_sort' % (rootdir, str(ws)[-6:-2])
    plt.savefig(filename)
    plt.clf()


for ws in wb.get_sheet_names():
    ws = wb[ws]
    N = ws['Q1'].value
    colnames = {'01names': [], '02concentration': [], '03nat_total': [], '04nat_1': [], '05nat_2': [], '06v/at': [],
                '07etotal': [], '08emix': [], '09eform': [], '10mmtotal': [], '11mmavg': [],
                '12mm1avg': [], '13mm2avg': [], '14ms_1': [], '15ms_2': [], '16ms_g': []}
    for entry in colnames:
        for cells in ws.iter_cols(min_col=int(entry[:2]), max_col=int(entry[:2]), max_row=N):
            for cell in cells:
                colnames[entry].append(cell.value)
    # initial_colors = [] ### colors of initial configuration
    # for entry in colnames:
    #     for cells in ws.iter_cols(min_col=int(entry[:2]), max_col=int(entry[:2]), max_row=N):
    #         for cell in cells:
    #             initial_colors.append(cell.fill.start_color.index)
    for system in range(len(colnames['01names'])):
        if colnames['12mm1avg'][system] < 0:
            colnames['11mmavg'][system] = colnames['11mmavg'][system] * (-1)
            colnames['12mm1avg'][system] = colnames['12mm1avg'][system] * (-1)
            colnames['13mm2avg'][system] = colnames['13mm2avg'][system] * (-1)
    border_colors = []
    markers = []
    concentrations = []
    for ms in colnames['16ms_g']:
        if ms == 'NM':
            markers.append('s')
        elif ms == 'FM':
            markers.append('^')
        elif ms == 'FiM':
            markers.append('D')
        else:
            markers.append('v')

    for concentration in colnames['02concentration']:
        if not (concentration in concentrations):
            concentrations.append(concentration)
    n_c = len(concentrations)

    pattern = re.compile('_[0-9]+')
    possible_fill_colors = []
    for i in range(29):
        possible_fill_colors.append([np.random.random(), np.random.random(), np.random.random(), 0.7])

    fill_colors = []
    for i in range(N):
        fill_colors.append(possible_fill_colors[int(re.search(pattern, colnames['01names'][i]).group(0)[1:])])

    states = [[], [], [], [], [], [], [], [], [], [], [], [], [], [],
              [], [], [], [], [], [], [], [], [], [], [], [], [], []]
    for i in range(N):
        states[int(re.search(pattern, colnames['01names'][i]).group(0)[1:]) - 1].append(colnames['07etotal'][i])
    while [] in states:
        states.pop(states.index([]))
    # print states
    gs = []
    non_gs = []
    stable_systems = []
    for i in states:
        # print i
        gs.append(colnames['07etotal'].index(min(i)))
        stable_systems.append(colnames['01names'][colnames['07etotal'].index(min(i))])
    open('%s/%s_stable.txt' % (rootdir, str(ws)[-6:-2]), 'w').write(str('\n'.join(stable_systems)))
    for i in range(N):
        if not(i in gs):
            non_gs.append(i)
    gs.sort(reverse=True)
    non_gs.sort(reverse=True)
    # print 'gs: ', gs
    # print 'non_gs: ', non_gs
    gs_colnames = copy.deepcopy(colnames)
    ngs_colnames = copy.deepcopy(colnames)
    for i in colnames:
        for j in non_gs:
            del gs_colnames[i][j]
        for k in gs:
            del ngs_colnames[i][k]

    # print gs_colnames
    # print ngs_colnames
    # break
    # for i in range(N):
    #     if i in gs:
    #         border_colors.append('blue')
    #     else:
    #         border_colors.append('green')
    # plot_emix()
    plot_eform()
    # plot_volume() #ok
    # break
    # plot_mmavg() #ok
    # print ws
    # plot_mmsortavg() #ok
    # plot_mmavg_and_mmsortavg() #not useful
    break
